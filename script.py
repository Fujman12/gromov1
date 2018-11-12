from openpyxl import load_workbook
from girls.models import Girl
import datetime
import time


def validate(date_text):
    try:
        print(date_text)
        time.strptime(str(date_text).split(' ')[0], '%Y-%m-%d')

    except ValueError:
        print(False)
        return False

    return True

def print_smth():

    wb = load_workbook(filename="GROMOV_MODEL_MANAGEMENT2.xlsx", read_only=True)

    ws = wb['GROMOV MODEL MANAGEMENT']

    for row in range(2, 130):
        print('Row: %s' % row)

        date_of_birth = ws.cell(row=row, column=6).value
        g = Girl(first_name=ws.cell(row=row, column=3).value,
                 last_name=ws.cell(row=row, column=4).value,
                 height=ws.cell(row=row, column=7).value,
                 parameters=ws.cell(row=row, column=8).value,
                 hair_color=ws.cell(row=row, column=10).value)

        if validate(date_of_birth):
            g.date_of_birth = date_of_birth

        if ws.cell(row=row, column=14).value is not None:
            if "http" in ws.cell(row=row, column=14).value:
                image_id = ws.cell(row=row, column=14).value.split('id=', 1)[1]
                g.image_id = image_id

        g.save()

print_smth()